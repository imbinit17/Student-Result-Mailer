from tkinter import *
import openpyxl as xl
from fpdf import FPDF
import sendEmail as UDM2
import random


def bestOf5(lang,lit,math,hin,phy,chem,bio,his,geo,com_eco):
    #GROUPING OF SUBJECTS
    group = [0,0,0,0,0]
    compulsoryEnglish = (int(lang) + int(lit)) / 2 #english
    group[0] = (int(phy) + int(chem) + int(bio)) / 3 #science
    group[1] = (int(his) + int(geo))/2 # arts
    group[2] = int(hin)
    group[3] = int(math)
    group[4] = int(com_eco)

    #BUBBLESORTING

    for i in range(0,len(group)-1):
        for j in range(len(group)-1):
            if(group[j]>group[j+1]):
                temp = group[j]  
                group[j] = group[j+1]  
                group[j+1] = temp


    #CALCULATING PERCENTAGE
    percentage = (group[4] + group[3] + group[2] + group[1] + compulsoryEnglish) / 5

    return percentage


def otpCheck():
    #otpByUser = " "
    otpByUser = str(otpEntry.get())
    if(str(otp)==otpByUser):
        window2.destroy()
        loginApproved()
    else:
        tempLabel = Label(window2,text="Wrong OTP",font=("Times New Roman",15)).pack()


def otpVerification():

    global window2
    window2 = Tk()
    window2.geometry("350x200")
    window2.title("OTP WINDOW")
    spaceLabel = Label(window2,text="",font=("Times New Roman",15)).pack()
    labelForOTP = Label(window2,text="Enter OTP : ",font=("Times New Roman",15)).pack()
    global otpEntry
    otpEntry = Entry(window2,width="15",font=("Times New Roman",14))
    otpEntry.pack()
    a = 1
    i = 1
    global otp
    while(i<=6):
        a = a + random.randint(0,9)

        a = a * 10
        i = i + 1
    otp = int(a/10)

    msg = "Your OTP for verification is : " + str(otp)

    mail = sheet.cell(row=index,column=15).value

    UDM2.sendMail(mail,"OTP",msg,[])
    tempLabel2 = Label(window2,text="").pack()

    submitButton = Button(window2,text="Submit",font=("Times New Roman",13),bg="Black",fg="White",command=otpCheck).pack()

    window2.mainloop()


def loginApproved():
    lang = sheet.cell(row=index,column=5).value
    lit = sheet.cell(row=index,column=6).value
    math = sheet.cell(row=index,column=7).value
    hin = sheet.cell(row=index,column=8).value
    phy = sheet.cell(row=index,column=9).value
    chem = sheet.cell(row=index,column=10).value
    bio = sheet.cell(row=index,column=11).value
    his = sheet.cell(row=index,column=12).value
    geo = sheet.cell(row=index,column=13).value
    com_eco = sheet.cell(row=index,column=14).value


    name = sheet.cell(row=index,column=3).value
    roll = sheet.cell(row=index,column=4).value
    email = sheet.cell(row=index,column=15).value
    mobile = sheet.cell(row=index,column=16).value
    
    percentage = bestOf5(lang,lit,math,hin,phy,chem,bio,his,geo,com_eco)

    txt1 = "NAME : " + str(name)
    txt2 = "ROLL : " + str(roll)
    txt3 = "EMAIL ID : "+str(email)
    txt4 = "MOBILE NO : "+ str(mobile)
    txtPer = "BEST OF 5 % :  " + str(percentage)

    txt5 = "English Language :  "+str(lang)
    txt6 = "English Literature :  " + str(lit)
    txt7 = "Maths :  "+str(math)
    txt8 = "Hindi :  "+ str(hin)
    txt9 = "Physics :  "+str(phy)
    txt10 = "Chemistry :  " + str(chem)
    txt11 = "Biology :  " + str(bio)
    txt12 = "History & Civics :  " + str(his)
    txt13 = "Geography :  " + str(geo)
    txt14 = "Computer/Economics Applications :  "+ str(com_eco)

    
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Times",'BIU',size = 35)
    pdf.cell(200,20,txt="Student Result",ln=1,align="C")
    pdf.set_font("Times",'I',size = 15)
    pdf.cell(200,10,txt=txt1,ln=2,align="L")
    pdf.cell(200,10,txt=txt2,ln=3,align="L")
    pdf.cell(200,10,txt=txt3,ln=4,align="L")
    pdf.cell(200,10,txt=txt4,ln=5,align="L")
    pdf.cell(200,10,txt=txtPer,ln=6,align="L")

    pdf.cell(200,10,txt="",ln=8,align="L")
    pdf.set_font("Times",size = 13)
    
    pdf.cell(200,10,txt=txt5,ln=7,align="L")
    pdf.cell(200,10,txt=txt5,ln=8,align="L")
    pdf.cell(200,10,txt=txt6,ln=9,align="L")
    pdf.cell(200,10,txt=txt7,ln=10,align="L")
    pdf.cell(200,10,txt=txt8,ln=11,align="L")
    pdf.cell(200,10,txt=txt9,ln=12,align="L")
    pdf.cell(200,10,txt=txt10,ln=13,align="L")
    pdf.cell(200,10,txt=txt11,ln=14,align="L")
    pdf.cell(200,10,txt=txt12,ln=15,align="L")
    pdf.cell(200,10,txt=txt13,ln=16,align="L")
    pdf.cell(200,10,txt=txt14,ln=17,align="L")

    pdf.cell(200,10,txt="",ln=18,align="L")
    pdf.cell(200,10,txt="",ln=19,align="L")
    pdf.cell(200,10,txt="***This is a system generated auto result and hence requires no signature ",ln=20,align="L")
    

    pdf.output("Result.pdf")

    message = '''The attachment containing the result is herewith the email .
Thank You !    '''

    UDM2.sendMail(email,"Result",message,['Result.pdf'])

    msg = "Login Granted ! Kindly Check Your Email ."
    msgLabel = Label(window,text=msg,fg="Green",font=("Cambria",12))
    msgLabel.place(x=50,y=90)


def login():
    userID = str(userIDEntry.get())
    pw = str(pwEntry.get())

    global workbook
    global sheet
    workbook = xl.load_workbook("Data.xlsx")
    sheet = workbook.active
    i = 1
    global index
    index = 0
    while(i<=sheet.max_row):
        if(sheet.cell(row = i,column=1).value==userID):
            index= index + i  
        i = i + 1

    if(index>0):
        if(sheet.cell(row=index,column = 2).value==pw):
            otpVerification()
        else:
            msg = "Invalid Password      "
            msgLabel = Label(window,text=msg,fg="Red",font=("Cambria",12))
            msgLabel.place(x=110,y=90)
    else :
        msg = "User doesn't exist     "
        msgLabel = Label(window,text=msg,fg="Red",font=("Cambria",12))
        msgLabel.place(x=110,y=90)
            


def func():
    global window
    window = Tk()
    window.geometry("350x170")
    window.title("Student Login Dashboard")
    window.config(bg="Sky Blue")

    userLabel = Label(window,text="User ID :",font=("Times New Roman",13))
    userLabel.place(x=50,y=20)

    global userIDEntry
    userIDEntry = Entry(window,width="15",font=("Times New Roman",13))
    userIDEntry.place(x=135,y=20)

    pwdLabel = Label(window,text="Password :",font=("Times New Roman",13))
    pwdLabel.place(x=50,y=60)

    global pwEntry
    pwEntry = Entry(window,width="15",font=("Times New Roman",13))
    pwEntry.place(x=135,y=60)

    submitBtn = Button(window,text="Login",font=("Times New Roman",13),bg="Black",fg="White",command=login)
    submitBtn.place(x=150,y=120)

    window.mainloop()

func()
